"""
run_all_evals.py

Transcribes clean and distorted once, evaluates every restored_* directory,
writes a consolidated Excel workbook, and prints a summary table.

Usage
-----
    python3 evals/run_all_evals.py
    python3 evals/run_all_evals.py --output results/
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import eval_runner as er

DATA_ROOT = Path("data")

# ── Discover restored directories ─────────────────────────────────────────────

def discover_restored_dirs() -> list[tuple[str, Path]]:
    """Return [(label, path)] for each restored_* dir that contains WAVs.

    data/restored/ is included as label "restored" only if it is NOT a
    duplicate of restored_auto_any (same file set).
    """
    dirs: list[tuple[str, Path]] = []

    default_dir = DATA_ROOT / "restored"
    auto_any_dir = DATA_ROOT / "restored_auto_any"

    # Check if default restored/ is a duplicate of restored_auto_any
    default_is_dup = False
    if default_dir.exists() and auto_any_dir.exists():
        default_wavs = {f.name for f in default_dir.glob("*.wav")}
        auto_any_wavs = {f.name for f in auto_any_dir.glob("*.wav")}
        if default_wavs == auto_any_wavs:
            default_is_dup = True

    if not default_is_dup and default_dir.exists() and any(default_dir.glob("*.wav")):
        dirs.append(("restored", default_dir))

    for d in sorted(DATA_ROOT.glob("restored_*")):
        if d.is_dir() and any(d.glob("*.wav")):
            label = d.name[len("restored_"):]
            dirs.append((label, d))

    return dirs


# ── Terminal table ─────────────────────────────────────────────────────────────

def print_summary_table(rows: list[dict]) -> None:
    """Print a fixed-width table of condition results to stdout."""
    col_w = {"condition": 22, "n": 5, "correct": 8, "car": 8, "wer": 9}
    header = (
        f"{'Condition':<{col_w['condition']}}"
        f"{'N':>{col_w['n']}}"
        f"{'Correct':>{col_w['correct']}}"
        f"{'CAR':>{col_w['car']}}"
        f"{'Mean WER':>{col_w['wer']}}"
    )
    sep = "-" * len(header)
    print("\n" + sep)
    print(header)
    print(sep)
    for r in rows:
        wer_str = f"{r['mean_wer']:.3f}" if r["mean_wer"] is not None else "  n/a"
        print(
            f"{r['condition']:<{col_w['condition']}}"
            f"{r['n']:>{col_w['n']}}"
            f"{r['correct']:>{col_w['correct']}}"
            f"{r['car']:>{col_w['car']}.1%}"
            f"{wer_str:>{col_w['wer']}}"
        )
    print(sep + "\n")


# ── Excel output ───────────────────────────────────────────────────────────────

def write_excel(
    summary_rows: list[dict],
    all_raw: list[dict],
    output_path: Path,
) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2F5597")
    clean_fill = PatternFill("solid", fgColor="D9EAD3")
    dist_fill  = PatternFill("solid", fgColor="FCE5CD")
    rest_fill  = PatternFill("solid", fgColor="CFE2F3")

    sum_headers = ["Condition", "N", "Correct", "CAR", "Mean WER"]
    ws_sum.append(sum_headers)
    for col, _ in enumerate(sum_headers, 1):
        cell = ws_sum.cell(1, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for r in summary_rows:
        wer_val = r["mean_wer"] if r["mean_wer"] is not None else ""
        ws_sum.append([r["condition"], r["n"], r["correct"], r["car"], wer_val])
        row_idx = ws_sum.max_row
        cond = r["condition"]
        if cond == "clean":
            fill = clean_fill
        elif cond == "distorted":
            fill = dist_fill
        else:
            fill = rest_fill
        for col in range(1, 6):
            ws_sum.cell(row_idx, col).fill = fill

        # Format CAR as percentage
        ws_sum.cell(row_idx, 4).number_format = "0.0%"
        if wer_val != "":
            ws_sum.cell(row_idx, 5).number_format = "0.000"

    # Column widths
    for col, width in zip(range(1, 6), [24, 6, 9, 9, 10]):
        ws_sum.column_dimensions[get_column_letter(col)].width = width

    # ── Raw sheet ──────────────────────────────────────────────────────────
    ws_raw = wb.create_sheet("Raw")
    raw_headers = ["condition", "file", "action_id", "ground_truth",
                   "predicted", "correct", "wer", "transcript"]
    ws_raw.append(raw_headers)
    for col in range(1, len(raw_headers) + 1):
        cell = ws_raw.cell(1, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for row in all_raw:
        ws_raw.append([row.get(h, "") for h in raw_headers])

    # Auto-width for raw sheet (cap at 60)
    for col_cells in ws_raw.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws_raw.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

    wb.save(output_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Run all eval conditions and write consolidated Excel results."
    )
    parser.add_argument("--output", default="results/",
                        help="Output directory (default: results/).")
    args = parser.parse_args()

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    restored_dirs = discover_restored_dirs()
    if not restored_dirs:
        print("[ERROR] No restored_* directories with WAV files found under data/")
        sys.exit(1)

    print(f"Found {len(restored_dirs)} restored condition(s):")
    for label, path in restored_dirs:
        print(f"  {label:25s}  {path}")

    # ── Step 1: Transcribe clean to get WER references ──────────────────
    clean_dir = er.DATA_DIRS["clean"]
    reference_transcripts: dict[str, str] = {}
    if clean_dir.exists():
        print("\nStep 1/3 — Transcribing clean audio for WER references …")
        reference_transcripts = er.build_reference_transcripts(clean_dir)
        print(f"  {len(reference_transcripts)} references ready.")
    else:
        print("[WARN] data/clean/ not found — WER will be skipped.")

    # ── Step 2: Run clean + distorted ────────────────────────────────────
    all_raw: list[dict] = []
    summary_rows: list[dict] = []

    for cond in ("clean", "distorted"):
        data_dir = er.DATA_DIRS[cond]
        if not data_dir.exists():
            print(f"\n[SKIP] {cond}: {data_dir} not found")
            continue
        print(f"\nStep 2/3 — Running condition: {cond} ({data_dir})")
        refs = reference_transcripts if cond != "clean" else None
        results = er.run_condition(cond, data_dir, reference_transcripts=refs)
        if cond == "clean" and not reference_transcripts:
            reference_transcripts = {Path(r["file"]).stem: r["transcript"] for r in results}
        all_raw.extend(results)
        n = len(results)
        correct = sum(r["correct"] for r in results)
        wer_vals = [r["wer"] for r in results if r.get("wer") is not None]
        summary_rows.append({
            "condition": cond,
            "n": n,
            "correct": correct,
            "car": round(correct / n, 4) if n else 0.0,
            "mean_wer": round(sum(wer_vals) / len(wer_vals), 4) if wer_vals else None,
        })
        print(f"  {n} samples evaluated")

    # ── Step 3: Run each restored condition ───────────────────────────────
    print(f"\nStep 3/3 — Running {len(restored_dirs)} restored condition(s) …")
    for label, path in restored_dirs:
        print(f"\n  Condition: {label} ({path})")
        results = er.run_condition(label, path, reference_transcripts=reference_transcripts)
        all_raw.extend(results)
        n = len(results)
        correct = sum(r["correct"] for r in results)
        wer_vals = [r["wer"] for r in results if r.get("wer") is not None]
        summary_rows.append({
            "condition": label,
            "n": n,
            "correct": correct,
            "car": round(correct / n, 4) if n else 0.0,
            "mean_wer": round(sum(wer_vals) / len(wer_vals), 4) if wer_vals else None,
        })
        print(f"    {n} samples evaluated")

    # ── Print terminal table ──────────────────────────────────────────────
    print_summary_table(summary_rows)

    # ── Write outputs ─────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    raw_path = output_dir / f"all_raw_{timestamp}.csv"
    fieldnames = ["condition", "file", "action_id", "ground_truth",
                  "predicted", "correct", "wer", "transcript"]
    with open(raw_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_raw)

    xlsx_path = output_dir / f"all_results_{timestamp}.xlsx"
    write_excel(summary_rows, all_raw, xlsx_path)

    print(f"Raw CSV  -> {raw_path}")
    print(f"Excel    -> {xlsx_path}")


if __name__ == "__main__":
    main()
